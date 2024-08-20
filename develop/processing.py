import xml.etree.ElementTree as ET
import openpyxl as xls


def proc_mallas(data_path, output_path):
    # activate excel workbook and properties
    wb = xls.Workbook()
    sheet = wb.active
    rowJobName = 1
    rowData = 2
    column_key = 1
    column_value = 2

    # reading xml file
    print('Abriendo archivo xml')
    tree = ET.parse(data_path)
    root = tree.getroot()

    # processing xml file
    print('Procesando información xml --> -->')
    for folder in root:
        for key, value in folder.attrib.items():
            if key == 'FOLDER_NAME':
                print(key, " ", value)
                sheet.cell(rowJobName, column_key, key)
                sheet.cell(rowJobName, column_value, value)
        for jobs in folder:
            for key, value in jobs.attrib.items():
                if key == 'JOBNAME':
                    print(key, " ", value)
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1
            print("-----------------")
            sheet.cell(rowData, column_key, "-----")
            sheet.cell(rowData, column_value, "-----")
            # column_key += 2
            # column_value += 2

    wb.save(output_path)
    print('** Archivo excel creado con exito **')
